import pandas as pd
import sqlite3
import openpyxl
import csv

#use /home/pi/USDividendChampions1.xlsx on the raspberry pi

workbook = openpyxl.load_workbook(r'c:\users\nicka\pycharmprojects\raspberrypiscripts\usdividendchampions.xlsx')
if "Champions" in workbook.sheetnames:
    sheet = workbook["Champions"]
else:
    print("Error, unable to find champions sheet name = Champions")

with open('test.csv', 'w', newline="") as file:
    csv = csv.writer(file)
    sheet.delete_rows(1, 2)
    sheet.delete_cols(6, 33)
    sheet.delete_cols(3, 1)
    sheet.delete_cols(5, 1)
    for col in sheet.iter_cols():
        for row in sheet.iter_rows():
            # for cell in row:
                # print(cell.value)
            csv.writerow([cell.value for cell in row])

def insertdata(info):
    con = sqlite3.connect('dividendchampions.db')
    cur = con.cursor()
    indexnumber = 0
    for row in info:
        # stockinfo = info[indexnumber].tolist()
        yearsonlist = stockinfo.pop(3)
        cur.execute("insert into stockinfo (stockname, stickersymbol, sector, industry) values (?, ?, ?, ?)", stockinfo)
        cur.execute("insert into stockdates (dateofinfo, yearsonlist) values (?, ?)", ('curdate()', yearsonlist))
        indexnumber += 1
    con.commit()
    con.close()

# data_xls = pd.read_excel(r'c:\users\nicka\pycharmprojects\raspberrypiscripts\usdividendchampions.xlsx', sheet_name='Champions', engine='openpyxl', skiprows=2, index_col=0, usecols='A,B,D,E,AN')
#A = Symbol, B = Company, D = Sector, E = No Years, AN = Industry

# data_xls.to_csv('csvfile.csv', encoding='utf-8')
# data = pd.read_csv('csvfile.csv')

with open('test.csv', 'r') as file:
    for line in file:
        list = line.split(',')
        print(list)
        insertdata(list)




