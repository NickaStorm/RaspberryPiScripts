import pymysql
import openpyxl
import csv
import time
import mysql.connector

#python3 DividendExcel.py to run in Desktop directory

workbook = openpyxl.load_workbook(r'/home/ubuntu/RaspberryPiScripts/SQLdata/USDividendChampions.xlsx')
#workbook = openpyxl.load_workbook(r'c:\users\nicka\pycharmprojects\raspberrypiscripts\usdividendchampions.xlsx')

if "Champions" in workbook.sheetnames:
    sheet = workbook["Champions"]
else:
    print("Error, unable to find champions sheet name = Champions")

def convertXlToCsv():
    with open('test.csv', 'w', newline='') as file:
        tempcsv = csv.writer(file, delimiter = '%')
        #deletes the head rows that display names
        sheet.delete_rows(1, 3)
        #deletes everything that isn't necessary
        sheet.delete_cols(6, 33)
        sheet.delete_cols(3, 1)
        sheet.delete_cols(5, 1)
        for row in sheet.iter_rows():
            tempcsv.writerow([cell.value for cell in row])

def insertData(info):
    conn = mysql.connector.connect(
        host='localhost',
        user='user',
        password='blueberry',
        # port=3306,
        database='dividendchampions',
        auth_plugin='mysql_native_password'
    )

    cur = conn.cursor(prepared=True)
    yearsonlist = info.pop(3)

    stockinfo_query = """ INSERT INTO stockinfo
                       (stickersymbol, stockname, sector, industry) VALUES (%s,%s,%s,%s)"""
    stockdates_query = """ INSERT INTO stockdates
                           (dateofinfo, yearsonlist) VALUES (%s,%s)"""
    temptime = time.strftime('%Y-%m-%d')
    stockdates_info = (temptime, yearsonlist)

    cur.execute(stockinfo_query, info)
    cur.execute(stockdates_query, stockdates_info)

    conn.commit()
    conn.close()

def convertCsvPrep():
    with open('test.csv', 'r') as file:
        for line in file:
            list = line.split('%')
            list[4] = list[4].strip('\n')
            insertData(list)

convertXlToCsv()
convertCsvPrep()